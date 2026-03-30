"""
Экспорт откликов и вакансий в Excel (.xlsx).
Запускается вручную: python export_xlsx.py
"""

from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import db


STATUS_COLORS = {
    "applied":        "C6EFCE",  # зелёный
    "already_applied": "FFEB9C",  # жёлтый
    "pending":        "DDEBF7",  # голубой
    "error":          "FFC7CE",  # красный
    "skipped":        "F2F2F2",  # серый
}

COLUMNS = [
    ("Статус",           "status",       15),
    ("Дата отклика",     "applied_at",   18),
    ("Роль",             "role_alias",   10),
    ("Вакансия",         "title",        40),
    ("Компания",         "employer",     30),
    ("Зарплата от",      "salary_from",  14),
    ("Зарплата до",      "salary_to",    14),
    ("Валюта",           "salary_cur",   10),
    ("Город",            "area",         15),
    ("Ссылка",           "url",          50),
    ("Ошибка",           "error_msg",    30),
]


def _header_row(ws):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx, (header, _, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def export(output_path: str = "applications.xlsx"):
    db.init_db()
    rows = db.get_all_applications()

    wb = Workbook()
    ws = wb.active
    ws.title = "Отклики"

    _header_row(ws)

    for row_idx, app in enumerate(rows, start=2):
        status = app.get("status", "")
        fill_color = STATUS_COLORS.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=fill_color)

        for col_idx, (_, field, _) in enumerate(COLUMNS, start=1):
            value = app.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = fill
            if field == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"Сохранено: {output_path} ({len(rows)} строк)")


if __name__ == "__main__":
    export()
